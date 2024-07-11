import openpyxl

def find_empty_row(sheet, headers):
    def is_row_empty(sheet, row_index):
        for col_index in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row_index, column=col_index).value
            if cell_value:
                return False
        return True

    empty_row = None

    for i in range(1, sheet.max_row + 1):
        if is_row_empty(sheet, i):
            empty_row = i
            break

    if empty_row is None:
        empty_row = sheet.max_row + 1

    if empty_row == 1:
        for col_num, header in enumerate(headers):
            sheet.cell(row=1, column=col_num + 1, value=header)
        empty_row += 1

    return empty_row
